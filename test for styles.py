# -*- coding:utf-8 -*-
from openpyxl import *
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

wb = load_workbook(filename='testwb.xlsx')
ws = wb.active

print wb.ws.cell(row=1, cloumn=1).value
print wb.ws.cell(row=1, cloumn=1).border
print wb.ws.cell(row=1, cloumn=1).font
print wb.ws.cell(row=1, cloumn=1).alignment