# -*- coding:utf-8 -*-
import sys
from openpyxl import *
from openpyxl.styles import Font, Alignment, Side, Border
from openpyxl.chart import BarChart, Series, Reference ,LineChart
from openpyxl.chart.label import DataLabelList
reload(sys)
sys.setdefaultencoding('utf8')

path_install_data = '1.宽带与业务产品开通率统计数据（02W）.xlsx'
path_repair_data = '2.宽带与业务产品返修设备统计数据（02W）.xlsx'
# 读取开通数据，总共31个省份
wb_install_data = load_workbook(filename=path_install_data.decode('utf8'), data_only=True)
# 读取返修数据，总共31个省份
wb_repair_data = load_workbook(filename=path_repair_data.decode('utf8'), data_only=True)

outwb = Workbook()
datasheet_1 = outwb.create_sheet(u'各省份汇总', 0)
datasheet_2 = outwb.create_sheet(u'各省出货与开通数据', 1)
datasheet_3 = outwb.create_sheet(u'分省库存情况', 2)
datasheet_4 = outwb.create_sheet(u'返修', 3)
# outwb.remove_sheet(wb_install_data.get_sheet_by_name('Sheet'))

# 定义更新周次
week_num = 2
week_num_r = week_num-2

# 创建格式化样式
number_format_s = '0.00%'
font_s = Font(name='微软雅黑', size=9,)
alignment_s = Alignment(horizontal='center', vertical='center')
bd_s = Side(style='thin', color="000000")
border_s = Border(left=bd_s, top=bd_s, right=bd_s, bottom=bd_s)

# 针对每个省份合并单元格
for i in range(2, 311, 10):
    datasheet_1.merge_cells(start_row=i, start_column=1, end_row=i+9, end_column=1)
# 对合并的单元格写上每个省份
for i, j in zip(range(2, 311, 10), range(3, 34)):
    datasheet_1.cell(row=i, column=1).value = wb_install_data.sheetnames[j]

# 写第二列表头
for i in range(0, 31):
    datasheet_1.cell(row=2+10*i, column=2).value = "累计发货量"
    datasheet_1.cell(row=3+10*i, column=2).value = "累计开通量"
    datasheet_1.cell(row=4+10*i, column=2).value = "一级库存"
    datasheet_1.cell(row=5+10*i, column=2).value = "二级库存"
    datasheet_1.cell(row=6+10*i, column=2).value = "累计开通率"
    datasheet_1.cell(row=7+10*i, column=2).value = "周发货量"
    datasheet_1.cell(row=8+10*i, column=2).value = "周开通量"
    datasheet_1.cell(row=9+10*i, column=2).value = "当年发货量"
    datasheet_1.cell(row=10+10*i, column=2).value = "当年开通量"
    datasheet_1.cell(row=11+10*i, column=2).value = "开年开通率"

# 写首行周次
for i, j in zip(range(3, 3+week_num_r), range(7, 7+week_num_r)):
    _ws = wb_install_data.get_sheet_by_name(u"北京")
    datasheet_1.cell(row=1, column=i).value = _ws.cell(row=2, column=j).value

# 循环写每个省份每个周次数据
sheetnames = wb_install_data.get_sheet_names()
for sheet_i in range(3, 34):
    _ws = wb_install_data.get_sheet_by_name(sheetnames[sheet_i])

    for i, j in zip(range(3, 3+week_num_r), range(7, 7+week_num_r)):
        for l, m in zip(range(2+10*(sheet_i-3), 12+10*(sheet_i-3)), range(1733, 1743)):
            datasheet_1.cell(row=l, column=i).value = _ws.cell(row=m, column=j).value

# 江苏省数据特殊处理,江苏融合终端数据从1743行到1753行
ws = wb_install_data.get_sheet_by_name(u'江苏')
for i, j in zip(range(3, 3 + week_num_r), range(7, 7 + week_num_r)):
    for l, m in zip(range(2 + 10 * (10 - 3), 12 + 10 * (10 - 3)), range(1743, 1753)):
        datasheet_1.cell(row=l, column=i).value = ws.cell(row=m, column=j).value
# 湖北省数据特殊处理,湖北融合终端数据从1743行到1753行
ws = wb_install_data.get_sheet_by_name(u'湖北')
for i, j in zip(range(3, 3 + week_num_r), range(7, 7 + week_num_r)):
    for l, m in zip(range(2 + 10 * (21 - 3), 12 + 10 * (21 - 3)), range(1743, 1753)):
        datasheet_1.cell(row=l, column=i).value = ws.cell(row=m, column=j).value
# 重庆省数据特殊处理,江苏融合终端数据从1823行到1833行
ws = wb_install_data.get_sheet_by_name(u'重庆')
for i, j in zip(range(3, 3 + week_num_r), range(7, 7 + week_num_r)):
    for l, m in zip(range(2 + 10 * (26 - 3), 12 + 10 * (26 - 3)), range(1823, 1833)):
        datasheet_1.cell(row=l, column=i).value = ws.cell(row=m, column=j).value

# 全国数据整理
datasheet_1.merge_cells(start_row=313, start_column=1, end_row=313+9, end_column=1)
datasheet_1.cell(row=313, column=1).value = '总体'
datasheet_1.cell(row=313, column=2).value = "累计发货量"
datasheet_1.cell(row=314, column=2).value = "累计开通量"
datasheet_1.cell(row=315, column=2).value = "一级库存"
datasheet_1.cell(row=316, column=2).value = "二级库存"
datasheet_1.cell(row=317, column=2).value = "累计开通率"
datasheet_1.cell(row=318, column=2).value = "周发货量"
datasheet_1.cell(row=319, column=2).value = "周开通量"
datasheet_1.cell(row=320, column=2).value = "当年发货量"
datasheet_1.cell(row=321, column=2).value = "当年开通量"
datasheet_1.cell(row=322, column=2).value = "开年开通率"
# 循环写汇总数据，全国融合终端数据从1853行到1863行
ws = wb_install_data.get_sheet_by_name(u'全国')
for i, j in zip(range(3, 3 + week_num_r), range(7, 7 + week_num_r)):
    for l, m in zip(range(313, 323), range(1853, 1863)):
        datasheet_1.cell(row=l, column=i).value = ws.cell(row=m, column=j).value

# 格式化开通率数据为百分比
for _row in range(0, 31*2):
    for _col in range(3, 3 + week_num_r):
        datasheet_1.cell(row=6 + _row * 5, column=_col).number_format = number_format_s
# 对表1的全国汇总数据317行和322行单独处理百分比格式
for _col in range(3, 3 + week_num_r):
    datasheet_1.cell(row=317, column=_col).number_format = number_format_s
    datasheet_1.cell(row=322, column=_col).number_format = number_format_s

# 对表1格式化样式
for _row in range(1, 323):
    for _col in range(1, week_num_r+3):
        datasheet_1.cell(row=_row, column=_col).font = font_s
        datasheet_1.cell(row=_row, column=_col).alignment = alignment_s
        datasheet_1.cell(row=_row, column=_col).border = border_s


# 第二张表对首行写表头
datasheet_2.cell(row=1, column=1).value = '序号'
datasheet_2.cell(row=1, column=2).value = "省份"
datasheet_2.cell(row=1, column=3).value = "累计发货量"
datasheet_2.cell(row=1, column=4).value = "累计开通量"
datasheet_2.cell(row=1, column=5).value = "周发货量"
datasheet_2.cell(row=1, column=6).value = "周开通量"
datasheet_2.cell(row=1, column=7).value = "当年发货量"
datasheet_2.cell(row=1, column=8).value = "当年开通量"
# 第二张表写数据
for i in range(2, 34):
    datasheet_2.cell(row=i, column=1).value = i-1
for i, j in zip(range(2, 33), range(3, 34)):
    datasheet_2.cell(row=i, column=2).value = wb_install_data.sheetnames[j]
for i, j in zip(range(2, 34), range(0, 31)):
    datasheet_2.cell(row=i, column=3).value = datasheet_1.cell(row=2+j*10, column=week_num).value
for i, j in zip(range(2, 34), range(0, 31)):
    datasheet_2.cell(row=i, column=4).value = datasheet_1.cell(row=3+j*10, column=week_num).value
for i, j in zip(range(2, 34), range(0, 31)):
    datasheet_2.cell(row=i, column=5).value = datasheet_1.cell(row=7+j*10, column=week_num).value
for i, j in zip(range(2, 34), range(0, 31)):
    datasheet_2.cell(row=i, column=6).value = datasheet_1.cell(row=8+j*10, column=week_num).value
for i, j in zip(range(2, 34), range(0, 31)):
    datasheet_2.cell(row=i, column=7).value = datasheet_1.cell(row=9+j*10, column=week_num).value
for i, j in zip(range(2, 34), range(0, 31)):
    datasheet_2.cell(row=i, column=8).value = datasheet_1.cell(row=10+j*10, column=week_num).value
# 第二张表汇总数据填写
datasheet_2.cell(row=33, column=2).value = "总体情况"
datasheet_2.cell(row=33, column=3).value = datasheet_1.cell(row=313, column=week_num).value
datasheet_2.cell(row=33, column=4).value = datasheet_1.cell(row=314, column=week_num).value
datasheet_2.cell(row=33, column=5).value = datasheet_1.cell(row=318, column=week_num).value
datasheet_2.cell(row=33, column=6).value = datasheet_1.cell(row=319, column=week_num).value
datasheet_2.cell(row=33, column=7).value = datasheet_1.cell(row=320, column=week_num).value
datasheet_2.cell(row=33, column=8).value = datasheet_1.cell(row=321, column=week_num).value

# 对表2格式化样式
for _row in range(1, 34):
    for _col in range(1, 9):
        datasheet_2.cell(row=_row, column=_col).font = font_s
        datasheet_2.cell(row=_row, column=_col).alignment = alignment_s
        datasheet_2.cell(row=_row, column=_col).border = border_s

# 第三张表对首行写表头
datasheet_3.cell(row=1, column=1).value = "序号"
datasheet_3.cell(row=1, column=2).value = "省份"
datasheet_3.cell(row=1, column=3).value = "累计发货量"
datasheet_3.cell(row=1, column=4).value = "累计开通量"
datasheet_3.cell(row=1, column=5).value = "库存量"
datasheet_3.cell(row=1, column=6).value = "库存比例"

# 第三张表写数据
for i in range(2, 34):
    datasheet_3.cell(row=i, column=1).value = i-1
for i, j in zip(range(2, 33), range(3, 34)):
    datasheet_3.cell(row=i, column=2).value = wb_install_data.sheetnames[j]
for i, j in zip(range(2, 33), range(0, 31)):
    datasheet_3.cell(row=i, column=3).value = datasheet_1.cell(row=2+j*10, column=week_num).value
for i, j in zip(range(2, 33), range(0, 31)):
    datasheet_3.cell(row=i, column=4).value = datasheet_1.cell(row=3+j*10, column=week_num).value

for i in range(2, 33):
    datasheet_3.cell(row=i, column=5).value = datasheet_3.cell(row=i, column=3).value-datasheet_3.cell(row=i, column=4).value

for _row in range(2, 33):
    try:
        datasheet_3.cell(row=_row, column=6).value = float(datasheet_3.cell(row=_row, column=5).value)/float(datasheet_3.cell(row=_row, column=3).value)
    except ZeroDivisionError:
        datasheet_3.cell(row=_row, column=6).value = 0

#表3汇总数据填写
datasheet_3.cell(row=33, column=2).value = "总体情况"
datasheet_3.cell(row=33, column=3).value = datasheet_1.cell(row=313, column=week_num).value
datasheet_3.cell(row=33, column=4).value = datasheet_1.cell(row=314, column=week_num).value
datasheet_3.cell(row=33, column=5).value = datasheet_3.cell(row=33, column=3).value-datasheet_3.cell(row=33, column=4).value
datasheet_3.cell(row=33, column=6).value = float(datasheet_3.cell(row=33, column=5).value)/float(datasheet_3.cell(row=33, column=3).value)

# 对表3格式化样式
for _row in range(1, 34):
    for _col in range(1, 7):
        datasheet_3.cell(row=_row, column=_col).font = font_s
        datasheet_3.cell(row=_row, column=_col).alignment = alignment_s
        datasheet_3.cell(row=_row, column=_col).border = border_s
for _row in range(2, 34):
    datasheet_3.cell(row=_row, column=6).number_format = number_format_s

# 生成表2图表1
chart2_1 = BarChart()
chart2_1.style = 3
chart2_1.width= 30
chart2_1.title = "周开通量统计"
chart2_1.y_axis.title = '周开通量（台）'
data = Reference(datasheet_2, min_col=6, min_row=1, max_row=32)
cats = Reference(datasheet_2, min_col=2, min_row=2, max_row=32)
chart2_1.add_data(data, titles_from_data=True)
chart2_1.set_categories(cats)
datasheet_2.add_chart(chart2_1, "A35")
chart2_1.dataLabels = DataLabelList()
chart2_1.dataLabels.showVal = True
# 生成表2图表2
chart2_2 = BarChart()
chart2_2.style = 3
chart2_2.width= 30
chart2_2.title = "当年累计开通量统计"
chart2_2.y_axis.title = '当年累计开通量（台）'
data = Reference(datasheet_2, min_col=4, min_row=1, max_row=32)
cats = Reference(datasheet_2, min_col=2, min_row=2, max_row=32)
chart2_2.add_data(data, titles_from_data=True)
chart2_2.set_categories(cats)
datasheet_2.add_chart(chart2_2, "A52")
chart2_2.dataLabels = DataLabelList()
chart2_2.dataLabels.showVal = True
# 生成表3图表1
chart3_1 = BarChart()
chart3_1.style = 3
chart3_1.width= 30
chart3_1.title = "融合终端放装与库存情况"
data = Reference(datasheet_3, min_col=2, min_row=1, max_row=32, max_col=4)
cats = Reference(datasheet_3, min_col=2, min_row=2, max_row=32)
chart3_1.add_data(data, titles_from_data=True)
chart3_1.set_categories(cats)

# chart3_2 = LineChart()
# data = Reference(datasheet_3, min_col=6, min_row=2, max_row=32)
# chart3_2.add_data(data, titles_from_data=True, from_rows=True)
# chart3_2.y_axis.axId = 200
# chart3_2.y_axis.title = "库存比例"
# chart3_2.y_axis.crosses = "max"
# chart3_1 += chart3_2

datasheet_3.add_chart(chart3_1, "A35")
chart3_1.dataLabels = DataLabelList()
chart3_1.dataLabels.showVal = True

#返修率数据生成
datasheet_4.merge_cells(start_row=2, start_column=1, end_row=10, end_column=1)
datasheet_4.cell(row=2, column=1).value = "融合终端&商业终端"
datasheet_4.cell(row=2, column=2).value = "在网数量"
datasheet_4.cell(row=3, column=2).value = "周返修数量"
datasheet_4.cell(row=4, column=2).value = "周良品数量"
datasheet_4.cell(row=5, column=2).value = "周返修率"
datasheet_4.cell(row=6, column=2).value = "周良品率"
datasheet_4.cell(row=7, column=2).value = "当年返修数量"
datasheet_4.cell(row=8, column=2).value = "当年良品数量"
datasheet_4.cell(row=9, column=2).value = "当年返修率"
datasheet_4.cell(row=10, column=2).value = "当年良品率"

# 对表4格式化样式
for _row in range(1, 11):
    for _col in range(1, week_num_r+3):
        datasheet_4.cell(row=_row, column=_col).font = font_s
        datasheet_4.cell(row=_row, column=_col).alignment = alignment_s
        datasheet_4.cell(row=_row, column=_col).border = border_s
for _col in range(3, 100):
    for _row in [5,6,9,10]:
        datasheet_4.cell(row=_row, column=_col).number_format = number_format_s

# 表4写首行周次
_ws = wb_repair_data.get_sheet_by_name(u"全国")
for _col_1, _col_2 in zip(range(3, 3+week_num_r), range(7, 7+week_num_r)):
    datasheet_4.cell(row=1, column=_col_1).value = _ws.cell(row=2, column=_col_2).value
# 表4写数据
for _col_dst, _col_src in zip(range(3, 3 + week_num_r), range(7, 7 + week_num_r)):
    for _row_dst, _row_src in zip(range(2, 11), range(1623, 1632)):
        datasheet_4.cell(row=_row_dst, column=_col_dst).value = _ws.cell(row=_row_src, column=_col_src).value
# 生成表4图标
chart4_1 = BarChart()
data1 = Reference(datasheet_4, min_col=4, min_row=7, max_col=9)
cats1 = Reference(datasheet_4, min_col=4, min_row=1, max_col=9)
chart4_1.add_data(data1, titles_from_data=False, from_rows=True)
chart4_1.y_axis.majorGridlines = None
chart4_1.set_categories(cats1)
chart4_1.title = '融合终端返修情况'

chart4_2 = LineChart()
data2 = Reference(datasheet_4, min_col=4, min_row=9, max_col=9)
chart4_2.add_data(data2, titles_from_data=False, from_rows=True)
chart4_2.y_axis.axId = 200

chart4_2.y_axis.crosses = "max"
chart4_1 += chart4_2

chart4_2.dataLabels = DataLabelList()
chart4_2.dataLabels.showVal = True

datasheet_4.add_chart(chart4_1, "C12")


# 保存文件
outwb.save("datasheet1.xlsx")